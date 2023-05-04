# -*- coding: utf-8 -*-
"""
Created on Thu Mar  9 16:18:30 2023

@author: fabic
"""

import pandas as pd
import openpyxl

# Data sources
materials_data = pd.read_excel('roads_archetypes.xlsx', header = [0], sheet_name='Material_intensities')
area_data_values = pd.read_excel('roads_areas_for_RECC_reestimated.xlsx', header = [0], sheet_name='roads_areas_for_RECCreestimated')
area_data = openpyxl.load_workbook('roads_areas_for_RECC_reestimated.xlsx')
area_data_match = area_data['region_match']

# Data classification #
Cohort = range(1900, 2061)
Cohort2015 = range(1990, 2016)
Time = range(2015, 2061)

Scenarios = ['LED','SSP1','SSP2']

Sectors_infrastructures = { #RECC roads products and their correspondent GRIP index
    'Roads, highway': ('Highways', 1),
    'Roads, primary': ('Primary roads', 2),
    'Roads, secondary': ('Secondary roads', 3),
    'Roads, tertiary': ('Tertiary roads', 4),
    'Roads, local': ('Local roads', 5),
    }

Engineering_Materials_m2 = { # RECC materials and their correspondent name in the source data
    'asphalt': 'asphalt_int_median',
    'concrete aggregates': 'granular_int_median',
    'cement': 'cement_int_median',
    'concrete': 'concrete_int_median' 
    }

GRIP_region = { # GRIP regions and their correspondent index in the dataset
    'North America': 1,
    'Central and South America': 2,
    'Africa': 3,
    'Europe': 4,
    'Middle East and Central Asia': 5,
    'South and East Asia': 6,
    'Oceania': 7,
    'Global': 8
    }   

data_available_combinations = [ # available data in the dataset for material composition
    (1,  'USA'), 
    (1,	'CAN'),
    (1,	'generic'),
    
    (2,	'CHL'),
    (2,	'MEX'),
    (2,	'generic'),

    (3,	'ZAF'),
    (3,	'generic'),
    
    (4,	'DEU'),
    (4,	'GBR'),
    (4,	'ITA'),
    (4,	'ESP'),
    (4,	'AUT'),
    (4,	'generic'),
    
    (5,	'ARE'),
    (5,	'generic'),
    
    (6,	'IND'),
    (6,	'CHN'),
    (6,	'VNM'),
    (6,	'JPN'),
    (6,	'generic'),
    
    (7,	'AUS'),
    (7,	'NZL'),
    (7,	'generic'),
    
    (8,	'generic')
    ]

SSP_Regions_32 = { # RECC region and match-up with region in the dataset
    'France':         ('Europe', 'generic') ,
    'Germany':        ('Europe', 'DEU'),
    'Italy':          ('Europe', 'ITA'),
    'Spain':          ('Europe', 'ESP'),
    'UK':             ('Europe', 'GBR'),
    'Poland':         ('Europe', 'generic'),
    'R32USA':         ('North America', 'USA'),
    'R32CAN':         ('North America', 'CAN'),
    'R32CHN':         ('South and East Asia', 'CHN'), 
    'R32JPN':         ('South and East Asia', 'JPN'),
    'R32IND':         ('South and East Asia', 'IND'), 
    'Oth_R32EU12-H':  ('Europe', 'generic'),
    'R32EU12-M':      ('Europe', 'generic'),
    'Oth_R32EU15':    ('Europe', 'generic'),
    'R5.2OECD_Other': ('Global', 'generic'),
    'R5.2REF_Other':  ('Europe', 'generic'),
    'R5.2ASIA_Other': ('South and East Asia', 'generic'), 
    'R5.2LAM_Other':  ('Central and South America', 'generic'),
    'R5.2MNF_Other':  ('Middle East and Central Asia',	'generic'),
    'R5.2SSA_Other':  ('Africa', 'generic')
}

Infrastructures_lifetime = { # Assumption for roads lifetime
    'Roads, highway': 25,
    'Roads, primary': 25,
    'Roads, secondary': 25,
    'Roads, tertiary': 25,
    'Roads, local': 25,
    }


Infrastructures_maintenance = { # Assumption for relative maintenance requirements
    'Roads, highway': 0.03,
    'Roads, primary': 0.03,
    'Roads, secondary': 0.03,
    'Roads, tertiary': 0.03,
    'Roads, local': 0.03,
    }
# Expand by region
Infrastructures_maintenance_r = {}
for r in SSP_Regions_32:
    Infrastructures_maintenance_r[r] = Infrastructures_maintenance
    

area_region_dict = {} # Create match_up dictionary for area data
for r in SSP_Regions_32:
    area_region_dict[r] = []
    Rix = 2
    while area_data_match.cell(Rix,2).value != None:
        if area_data_match.cell(Rix,4).value == r:
            area_region_dict[r].append(area_data_match.cell(Rix,2).value)
        Rix+=1
        
road_network_extension = {} # Create dictionary road network extension. Unit is million m2
for r in SSP_Regions_32:
    road_network_extension[r] = {}
    for i in Sectors_infrastructures:
        road_network_extension[r][i]= area_data_values[ (area_data_values['country'].isin(area_region_dict[r])) &  (area_data_values['GP_RTP']==Sectors_infrastructures[i][1]) ]['area_mid_m2'].sum()/1e06
        
        
    


    




