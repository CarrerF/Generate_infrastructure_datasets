# -*- coding: utf-8 -*-
"""
Created on Fri Mar 10 15:05:09 2023

@author: fabic
"""
import pandas as pd
import openpyxl
from statistics import mean
from tqdm import tqdm
import numpy as np

import Infrastructures_classification as ifr

dataset_generation = { # SELECT THE DATASET TO GENERATE
    '3_MC_RECC_Infrastructures':                     False,
    '2_S_RECC_FinalProducts_2015_infrastructures':   True,
    '3_LT_RECC_ProductLifetime_infrastructures':     False,
    '3_MC_RECC_Buildings_Renovation_Relative':       False,
    '2_S_RECC_FinalProducts_Future_infrastructures': True
    }

data = ifr.materials_data

if dataset_generation['3_MC_RECC_Infrastructures'] == True:
    # Generate a TABLE for infrastructure material composition
    # ROWS: Sectors_infrastructures, SSP_Regions_32
    # COLUMNS: Engineering_Materials_m2
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    sheet.cell(1,1).value = 'i'
    sheet.cell(1,2).value = 'r'
    Rix = 1 # Row index
    Cix = 3 # Column index
    for m in ifr.Engineering_Materials_m2:
        sheet.cell(Rix,Cix).value = m
        Cix +=1
    # Loop over columns and rows
    Rix = 2 
    for r in ifr.SSP_Regions_32:
        for i in ifr.Sectors_infrastructures:
            sheet.cell(Rix,1).value = i
            sheet.cell(Rix,2).value = r
            Cix = 3
            for m in ifr.Engineering_Materials_m2:
                    sheet.cell(Rix,Cix).value = mean(ifr.materials_data[(
                            ifr.materials_data['GRIP region'] == ifr.GRIP_region[ifr.SSP_Regions_32[r][0]]) 
                            & (ifr.materials_data['Country Alpha-3 Code'] == ifr.SSP_Regions_32[r][1])
                            & (ifr.materials_data['GRIP road type'] ==  ifr.Sectors_infrastructures[i][1])  ][ifr.Engineering_Materials_m2[m]] )
                    Cix+=1
            Rix+=1
    dataset.save('3_MC_RECC_Infrastructures.xlsx')
    print('3_MC_RECC_Infrastructures generated!')               

if dataset_generation['3_LT_RECC_ProductLifetime_infrastructures'] == True:
    # Generate a LIST for infrastructure lifetime
    # COLUMNS: Sectors_infrastructures, SSP_Regions_32
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    sheet.cell(1,1).value = 'i'
    sheet.cell(1,2).value = 'r'
    sheet.cell(1,3).value = 'Value'
    sheet.cell(1,4).value = 'Unit'
    # Loop over columns and rows
    Rix = 2 #Row index
    for r in ifr.SSP_Regions_32:
        for i in ifr.Sectors_infrastructures:
            sheet.cell(Rix,1).value = i
            sheet.cell(Rix,2).value = r
            sheet.cell(Rix,3).value = ifr.Infrastructures_lifetime[i]
            sheet.cell(Rix,4).value = 'yr'
            Rix +=1
    dataset.save('3_LT_RECC_ProductLifetime_infrastructures.xlsx')
    print('3_LT_RECC_ProductLifetime_infrastructures generated!')   
            
    
if dataset_generation['3_MC_RECC_Buildings_Renovation_Relative'] == True:
    # Generate a TABLE for infrastructure renovation
    # ROWS: Time, Engineering_Materials_m2 
    # COLUMNS: Sectors_infrastructures, SSP_Regions_32
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    # Loop over columns and rows
    Rix = 3 #Row index
    for t in ifr.Time:
        for m in ifr.Engineering_Materials_m2:
            Cix = 3 #Row index
            for r in ifr.SSP_Regions_32:
                for i in ifr.Sectors_infrastructures: 
                    sheet.cell(Rix,1).value = t
                    sheet.cell(Rix,2).value = m
                    sheet.cell(1,Cix).value = i
                    sheet.cell(2,Cix).value = r
                    sheet.cell(Rix,Cix).value = ifr.Infrastructures_maintenance_r[r][i]
                    Cix+=1
            Rix+=1
    dataset.save('3_MC_RECC_Infrastructures_Maintenance_Relative.xlsx')
    print('3_MC_RECC_Buildings_Maintenance_Relative generated!')   
    
    
if dataset_generation['2_S_RECC_FinalProducts_2015_infrastructures'] == True:
    # Generate a TABLE for infrastructure stock in 2015
    # ROWS: Time, Cohort
    # COLUMNS: Sectors_infrastructures, SSP_Regions_32
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    # Loop over columns and rows
    Rix = 3 #Row index
    for c in ifr.Cohort2015:
        Cix = 3 #Row index
        for r in ifr.SSP_Regions_32:
            for i in ifr.Sectors_infrastructures: 
                sheet.cell(Rix,1).value = '2015'
                sheet.cell(Rix,2).value = c
                sheet.cell(1,Cix).value = i
                sheet.cell(2,Cix).value = r
                if r in ['France','Germany','Italy','Spain','UK','Poland','R32USA','R32CAN','R32JPN']:
                    scale =1/len(ifr.Cohort2015) # model homogenous distribution
                if r in ['Oth_R32EU12-H','R32EU12-M','Oth_R32EU15','R32CHN','R32IND','R5.2OECD_Other','R5.2REF_Other','R5.2ASIA_Other','R5.2LAM_Other','R5.2MNF_Other','R5.2SSA_Other']:
                    scale =  2 / (len(ifr.Cohort2015)) / (len(ifr.Cohort2015)+1) * (ifr.Cohort2015.index(c)+1)
                sheet.cell(Rix,Cix).value = ifr.road_network_extension[r][i] *scale
                Cix+=1
        Rix+=1
    # Add a sheet with total network extension.
    # ROWS: SSP_Regions_32
    # COLUMNS: total   
    dataset.create_sheet('Total')
    sheet2 = dataset['Total']
    sheet2.cell(1,2).value = 'Total'
    Rix=2
    for r in ifr.SSP_Regions_32:
        sheet2.cell(Rix,1).value = r
        sheet2.cell(Rix,2).value = sum(ifr.road_network_extension[r].values())
        Rix+=1
    # Add a sheet with initial type split.
    # ROWS: SSP_Regions_32
    # COLUMNS: Sectors_infrastructures
    dataset.create_sheet('TypeSplit2015')
    sheet2 = dataset['TypeSplit2015']
    Rix = 1 
    Cix = 2
    for i in ifr.Sectors_infrastructures:
        sheet2.cell(Rix,Cix).value = i
        Cix+=1
    Rix = 2    
    for r in ifr.SSP_Regions_32:
        Cix = 2
        for i in ifr.Sectors_infrastructures:
            sheet2.cell(Rix,1).value = r
            sheet2.cell(Rix,Cix).value = ifr.road_network_extension[r][i] / sum(ifr.road_network_extension[r].values())
            Cix+=1
        Rix+=1
    dataset.save('2_S_RECC_FinalProducts_2015_infrastructures_tcir.xlsx')
    print('2_S_RECC_FinalProducts_2015_infrastructures_tcir generated!') 
    
    # Generate a TABLE for infrastructure stock in 2015
    # ROWS: Time
    # COLUMNS: Sectors_infrastructures, SSP_Regions_32
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    sheet.cell(1,1).value = 'r'
    sheet.cell(1,2).value = 'i'
    sheet.cell(1,3).value = '2015'
    Rix=2
    for r in ifr.SSP_Regions_32:
        for i in ifr.Sectors_infrastructures: 
            sheet.cell(Rix,1).value = r
            sheet.cell(Rix,2).value = i
            sheet.cell(Rix,3).value = ifr.road_network_extension[r][i]
            Rix+=1
    
    dataset.save('2_S_RECC_FinalProducts_2015_infrastructures_rit.xlsx')
    print('2_S_RECC_FinalProducts_2015_infrastructures_rit generated!') 
    
if dataset_generation['2_S_RECC_FinalProducts_Future_infrastructures'] == True:
    # Placeholder
    # Generate a TABLE for infrastructure stock in 2015
    # ROWS: Region, Infrastructure, Time
    # COLUMNS: Scenario
    dataset = openpyxl.Workbook()
    sheet = dataset.active
    sheet.cell(1,1).value = 'r'
    sheet.cell(1,2).value = 'i'
    sheet.cell(1,3).value = 't'
    Cix = 4
    for s in ifr.Scenarios:
        sheet.cell(1,Cix).value = s
        Cix +=1
    Rix = 2
    for r in ifr.SSP_Regions_32:
        for i in ifr.Sectors_infrastructures:
            for t in ifr.Time:
                Cix = 4
                for s in ifr.Scenarios:
                    sheet.cell(Rix,1).value = r
                    sheet.cell(Rix,2).value = i
                    sheet.cell(Rix,3).value = t
                    sheet.cell(Rix,Cix).value = ifr.road_network_extension[r][i]
                    Cix+=1
                Rix+=1
    dataset.save('2_S_RECC_FinalProducts_Future_infrastructures.xlsx')
    print('2_S_RECC_FinalProducts_Future_infrastructures generated!')             
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    